from datetime import date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

START = date(2026, 9, 3)
RES = {
    'Python': ('廖雪峰 Python 教程', 'https://liaoxuefeng.com/books/python/introduction/index.html'),
    'LLM': ('Datawhale Happy-LLM', 'https://github.com/datawhalechina/happy-llm'),
    'Agent': ('Datawhale Hello-Agents', 'https://github.com/datawhalechina/hello-agents'),
    'Ollama': ('Ollama 文档', 'https://ollama.readthedocs.io/quickstart/'),
    'LangGraph': ('LangGraph 文档', 'https://langchain-ai.github.io/langgraph/'),
    'Git': ('廖雪峰 Git 教程', 'https://liaoxuefeng.com/books/git/introduction/index.html'),
}
WEEKS = [
('Python 基础与工具链', 'Python', ['变量、字符串、输入输出|自我介绍生成器', '条件判断与循环|成绩分级与猜数字', '列表、元组、字典、集合|文本词频统计', '函数与模块|将词频程序拆成函数', '文件读写与 JSON|读写学习日志 JSON', 'Git 与虚拟环境|建立仓库并提交代码', '周测与复盘|不看教程重写小工具']),
('Python 进阶与数据处理', 'Python', ['异常处理与调试|为文件程序加入异常提示', '面向对象基础|实现 Note 与 Task 类', '推导式与生成器|批量处理文本文件', '正则与文本清洗|清理课程笔记格式', 'HTTP、JSON 与 API 概念|调用公开 API 并保存', '综合：命令行学习助手|实现查询、记录、统计', '周测与复盘|整理 README 与问题清单']),
('大模型原理', 'LLM', ['NLP 任务与语言模型|画文本到预测流程', 'Token 与分词|比较 Token 与字符差异', 'Embedding 与向量相似度|计算余弦相似度', 'Transformer 总览|画 Encoder/Decoder 数据流', 'Attention 注意力机制|手算小矩阵例子', '预训练、微调、推理|写一页对比笔记', '案例复现与周测|运行章节代码并解释输出']),
('LLM 应用开发', 'Agent', ['Prompt 结构与角色|同一任务写 3 版 Prompt', '结构化输出 JSON|课程信息抽取器', '多轮对话与上下文|保存并截断对话历史', 'Function/Tool Calling|接入计算器工具', '模型评测与幻觉|设计 10 条测试集', '综合：学习计划生成器|输出周计划 JSON', '周测与复盘|截图展示可运行结果']),
('RAG 知识库', 'Agent', ['RAG 原理与适用场景|画 RAG 流程图', '文档解析与切分|切分 Markdown 笔记', 'Embedding 实操|生成向量并保存', '向量数据库与召回|实现 Top-K 检索', '重排、引用与答案约束|回答带来源段落', '综合：个人知识库 v1|导入 TXT/MD 并问答', '评测与复盘|建立 15 条问题准确率表']),
('Agent 核心', 'Agent', ['Agent 定义、类型、范式|比较 Chatbot/Workflow/Agent', 'ReAct：思考-行动-观察|手写最小 ReAct 循环', '规划与任务分解|拆解做报告任务', '记忆与状态|实现短期记忆字典', '工具调用与权限|加文件读取工具并限制路径', '反思、重试与多 Agent|失败任务加一次重试', '案例复现与周测|完成单 Agent 学习助理']),
('框架与部署', 'LangGraph', ['LangChain/LangGraph 心智模型|跑通最小图工作流', '状态图、节点与边|实现检索到回答图', '低代码对照：Dify/Coze|复刻一个小流程', 'Ollama 本地模型安装|安装并运行小模型', 'OpenAI 兼容 API 与 FastAPI|暴露一个 chat 接口', 'Docker 基础部署|容器化并记录启动命令', '周测与复盘|画架构图与部署文档']),
('作品集与毕业项目', 'Agent', ['项目选题与需求文档|确定知识库或自动化 Agent', '项目架构与任务拆解|列模块、接口、风险', '项目开发 1：核心链路|完成主流程', '项目开发 2：工具与记忆|加入至少两个工具', '项目开发 3：评测与异常|加入测试集、日志、重试', '项目发布与展示|写 README、截图、演示视频', '最终复盘|总结 56 天成果与下一阶段']),
]

rows=[]
for wi, (module, default_res, lessons) in enumerate(WEEKS, 1):
    for di, lesson in enumerate(lessons, 1):
        topic, practice = lesson.split('|')
        res = 'Git' if topic.startswith('Git') else ('Ollama' if 'Ollama' in topic or 'Docker' in topic or 'FastAPI' in topic else default_res)
        d = START + timedelta(days=len(rows))
        weekdays = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
        rows.append([d.isoformat(), weekdays[d.weekday()], wi, module, topic, res, practice])

md=['# AI 与 Agent 56 天每日课程计划（2026-09-03 起）', '', '每天约 2 小时：15 分钟复习 + 30 分钟理论/案例 + 55 分钟实操 + 20 分钟记录。每周第 7 天复盘或周测。', '', '## 资源入口（优先免费）']
for key, (name, url) in RES.items(): md.append(f'- **{key}**：[ {name} ]({url})')
md += ['', '> 图片中的路线仅作参考。本计划优先让你形成可运行的项目成果；理论学到能解释、能应用即可。', '', '## 每日安排', '', '| 日期 | 星期 | 周次 | 模块 | 知识/案例 | 学习入口 | 今日实操产出 | 打卡 | 用时(分) | 备注 |', '|---|---|---:|---|---|---|---|---|---:|---|']
for d, weekday, week, module, topic, res, practice in rows:
    name,url=RES[res]
    md.append(f'| {d} | {weekday} | {week} | {module} | {topic} | [{name}]({url}) | {practice} | [ ] |  |  |')
md += ['', '## 成效时间', '- **2026-09-16（第 2 周末）**：能独立写并调试 Python 小工具。', '- **2026-09-30（第 4 周末）**：能调用模型、完成结构化输出的 LLM 小应用。', '- **2026-10-14（第 6 周末）**：能完成个人资料的基础 RAG 问答。', '- **2026-10-28（第 8 周末）**：能完成带工具调用、记忆和重试的单 Agent。', '- 后续 4 周可用来打磨部署和作品集项目。', '', '## 每周验收', '1. 至少完成 5 天；2. 至少提交 3 次代码；3. 周测作品可运行；4. 记录一个最难问题及解决方法。']
Path('AI_Agent_56天每日课程计划.md').write_text('\n'.join(md), encoding='utf-8')

wb=Workbook(); ws=wb.active; ws.title='每日打卡'
headers=['日期','星期','周次','模块','今日知识/案例','学习入口','今日实操产出','完成(√)','学习分钟','代码/笔记链接','今日困难','明日计划']
ws.append(headers)
for d, weekday, week, module, topic, res, practice in rows:
    ws.append([d,weekday,week,module,topic,RES[res][1],practice,'','','','',''])
for c in ws[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E78'); c.alignment=Alignment(horizontal='center',vertical='center')
for col,width in enumerate([14,10,8,20,28,54,35,12,12,26,26,26],1): ws.column_dimensions[chr(64+col)].width=width
for r in ws.iter_rows(min_row=2):
    for c in r: c.alignment=Alignment(vertical='top',wrap_text=True)
ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
summary=wb.create_sheet('每周复盘'); summary.append(['周次','日期范围','本周模块','完成天数','总分钟','本周作品/链接','最难问题','下周调整'])
for i,(module,_,_) in enumerate(WEEKS,1):
    s=START+timedelta(days=(i-1)*7); e=s+timedelta(days=6); summary.append([i,f'{s} ~ {e}',module,'','','','',''])
for c in summary[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='548235')
for col,width in zip('ABCDEFGH',[8,24,24,14,14,28,30,30]): summary.column_dimensions[col].width=width
resources=wb.create_sheet('资源清单'); resources.append(['分类','资源','链接','说明'])
for key,(name,url) in RES.items(): resources.append([key,name,url,'按每日计划使用'])
for c in resources[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='7F6000')
for col,width in zip('ABCD',[12,30,68,24]): resources.column_dimensions[col].width=width
wb.save('AI_Agent_56天打卡表.xlsx')
print('Created 56-day plan and workbook.')
